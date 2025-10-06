from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from .models import Employee, Skills
from .forms import EmployeeForm, SkillFormSet
from . import database_query
from Master import database_query as db
from Master.models import *

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
# logger = logging.getLogger(__name__)

# ------------------------------------------
# EMPLOYEE LIST
# ------------------------------------------
@login_required
def employee_list(request):
    try:
        username = request.user.username
        filters = {}
        if request.GET.get('date_from'):
            filters['date_from'] = request.GET.get('date_from')
        if request.GET.get('date_to'):
            filters['date_to'] = request.GET.get('date_to')
        rows = database_query.fetch_employees(filters)
        return render(request, 'employees/employee_list.html', {
            'employees': rows, 'username': username
        })
    except Exception as e:

        messages.error(request, f"Failed to load employee list: {e}")
        return render(request, 'employees/employee_list.html', {'employees': [], 'username': request.user.username})


# ------------------------------------------
# ADD EMPLOYEE
# ------------------------------------------
@login_required
def employee_add(request):
    username = request.user.username
    try:
        if request.method == 'POST':
            form = EmployeeForm(request.POST, request.FILES)
            formset = SkillFormSet(request.POST)
            if form.is_valid() and formset.is_valid():
                emp = form.save(commit=False)
                emp.created_by = username
                emp.updated_by = username
                emp.save()
                formset.instance = emp
                formset.save()
                messages.success(request, 'Employee created successfully.')
                return redirect('employee_list')
            else:
                messages.warning(request, 'Please correct the form errors.')
        else:
            form = EmployeeForm()
            formset = SkillFormSet()

        departments = db.get_departments()
        locations = db.get_locations()
        return render(request, 'employees/employee_form.html', {
            'form': form,
            'formset': formset,
            'departments': departments,
            'locations':locations,
            'username': username,
            'action': 'Add'
        })
    except Exception as e:
        messages.error(request, f"Error while adding employee: {e}")
        return redirect('employee_list')


# ------------------------------------------
# EDIT EMPLOYEE
# ------------------------------------------
@login_required
def employee_edit(request, pk):
    username = request.user.username
    try:
        emp = get_object_or_404(Employee, pk=pk)
        if request.method == 'POST':
            form = EmployeeForm(request.POST, request.FILES, instance=emp)
            formset = SkillFormSet(request.POST, instance=emp)
            if form.is_valid() and formset.is_valid():
                emp = form.save(commit=False)
                emp.updated_by = username
                emp.save()
                formset.save()
                messages.success(request, 'Employee updated successfully.')
                return redirect('employee_list')
            else:
                messages.warning(request, 'Please correct the form errors.')
        else:
            form = EmployeeForm(instance=emp)
            formset = SkillFormSet(instance=emp)

        departments = db.get_departments()
        locations = db.get_locations()
        return render(request, 'employees/employee_form.html', {
            'form': form,
            'formset': formset,
            'departments': departments,
            'locations':locations,
            'username': username,
            'action': 'Edit'
        })
    except Exception as e:
        messages.error(request, f"Error editing employee: {e}")
        return redirect('employee_list')


# ------------------------------------------
# AJAX: DESIGNATIONS BY DEPARTMENT
# ------------------------------------------
@login_required
# def ajax_designations(request):
#     try:
#         dept_id = request.GET.get('dept_id')
#         data = db.get_designations_by_department(dept_id)
#         print("data: ",data)
#         return JsonResponse({'designations': data})
#     except Exception as e:
#         logger.error(f"Error in ajax_designations: {e}")
#         return JsonResponse({'error': str(e)}, status=500)  
@login_required
def ajax_designations(request):
    dept_id = request.GET.get('dept_id')
    print("Selected Department ID:", dept_id)

    if not dept_id:
        return JsonResponse({'designations': []})

    designations = Designation.objects.filter(department_id=dept_id).values(
        'designation_id', 'des_name'
    )

    return JsonResponse({'designations': list(designations)})

# ------------------------------------------
# EMPLOYEE DETAIL
# ------------------------------------------
@login_required
def employee_detail(request, pk):
    username = request.user.username
    emp = get_object_or_404(Employee, pk=pk)
    skills = emp.skills.all().order_by('skills_name')  
    print(f'emp -- {emp},{skills}')
    return render(request, 'employees/employee_detail.html', {
        'emp': emp,
        'skills': skills,
        'username': username
    })
    # try:
    #     username = request.user.username
    #     emp = get_object_or_404(Employee, pk=pk)
    #     skills = emp.skills.all().order_by('skills_name')  
    #     print(f'emp -- {emp},{skills}')
    #     return render(request, 'employees/employee_detail.html', {
    #         'emp': emp,
    #         'skills': skills,
    #         'username': username
    #     })
    # except Exception as e:
    #     logger.error(f"Error in employee_detail: {e}")
    #     messages.error(request, f"Unable to load employee details: {e}")
    #     return redirect('employee_list')


# ------------------------------------------
# DELETE EMPLOYEE
# ------------------------------------------
@login_required
def employee_delete(request, pk):
    try:
        username = request.user.username
        emp = get_object_or_404(Employee, pk=pk)
        if request.method == 'POST':
            emp.delete()
            messages.success(request, 'Employee deleted successfully.')
            return redirect('employee_list')
        return render(request, 'employees/employee_detail.html', {'object': emp, 'username': username})
    except Exception as e:
        messages.error(request, f"Failed to delete employee: {e}")
        return redirect('employee_list')


# ------------------------------------------
# DOWNLOAD SELECTED EMPLOYEES (excel)
# ------------------------------------------

def download_selected(request):
    # Create workbook & sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Header row
    headers = [
        "Employee No", "Name", "Phone", "Department",
        "Designation", "Location", "Status", "Join Date","Start Date","Skills","Photo"
    ]
    ws.append(headers)

    # Query employees
    employees = Employee.objects.select_related("department", "designation", "location").all()

    # Add rows (convert related model objects to strings)
    for emp in employees:
        ws.append([
            emp.empno,
            emp.name,
            emp.phone or "",
            str(emp.department) if emp.department else "",
            str(emp.designation) if emp.designation else "",
            str(emp.location) if emp.location else "",
            emp.status.capitalize(),
            emp.join_date.strftime("%Y-%m-%d") if emp.join_date else "",
            emp.emp_start_date.strftime("%Y-%m-%d") if emp.join_date else "",
            ", ".join(emp.skills.values_list('skills_name', flat=True)),
            emp.photo.url if emp.photo else '',
        ])

    #  Set all column widths to 18
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    #  Freeze header row
    ws.freeze_panes = "A2"

    #  Style header row
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    #  Return Excel file as HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="employees.xlsx"'

    wb.save(response)
    return response
